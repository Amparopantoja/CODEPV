#Importar las librerias
import pandas as pd
import os
from openpyxl import load_workbook

#Crear variables y clase de PVF (Punto de venta; Factura; Cliente; Fecha)
file_name = "Clientes_a_Controlar 20.07.2023.xlsx"
sheet_name = "Hoja1"
carpeta_actual = os.path.dirname(os.path.abspath(__file__))
carpeta_actual = carpeta_actual.split('\\')[-1]
mes_actual = carpeta_actual.split('-')[0]
año_actual = int(carpeta_actual.split('-')[1])
año_mes = f'{año_actual}_{mes_actual}'


class PVF:
    def __init__(self, cliente, factura, fecha, punto_de_venta, n_factura):
        self.cliente = cliente
        self.factura = factura
        self.fecha = fecha
        self.punto_de_venta = punto_de_venta
        self.n_factura = n_factura 

class Cliente:
    def __init__(self, id_cliente, razon_social, cuit, linea):
        self.id_cliente = id_cliente
        self.razon_social = razon_social
        self.cuit = cuit
        self.linea = linea

#Abrir el excel Clientes a controlar (Pandas y openpy)
pd_doc = pd.read_excel(file_name, sheet_name)
openpy_doc = load_workbook(file_name)
nombres_hojas = openpy_doc.sheetnames
if 'Hoja1' in nombres_hojas:
    hoja_base = openpy_doc['Hoja1']
    
    #Asignación de formato a las columnas
    #nombre_columna_1 = pd_doc.iloc[:0]
    #nombre_columna_2 = pd_doc.iloc[:1]
    #nombre_columna_3 = pd_doc.iloc[:4]

    #pd_doc.columns[1] = pd_doc.columns[1].astype(str)
    #pd_doc.columns[2] = pd_doc.columns[2].astype(str)
    #pd_doc.columns[3] = pd_doc.columns[3].astype(str)
    
    #pd_doc = pd_doc.astype(str)
    #Obtener la lista de clientes
    lista_clientes = []


    #Recorre las lineas del excel con pandas. i = numero de iteración; f = linea de cliente en el excel
    for i, f in pd_doc.iterrows():
        if(i > 0):
            #Creación de instancia de Cliente
            cc = Cliente(f[0], f[1], f[3], i+2)
            #Lo agrega a la lista de contribuyentes
            lista_clientes.append(cc)

    index = 0

    #Acceder a la carpeta de cada cliente (For)
    for c in lista_clientes:
        index +=1

        #Armar ruta(path)para las carpetas de los clientes    
        try:
            excel_path = ''
            path_folder = '..\..\XLS -clientes'
            folders = os.listdir(path_folder)

            exists_folder = False
            for f in folders:
                if f == f'{c.id_cliente}_{c.cuit}':
                    excel_path = str(f'{path_folder}/{f}/{año_mes}')
                    exists_folder = True
                    break 
            
            if exists_folder:    
                excel_files = os.listdir(excel_path)

                exists_excel = False
                for e in excel_files:
                    if "Mis Comprobantes Emitidos" in e:
                        exists_excel = True
                        #Abra el excel Mis comprobantes emitidos (Pandas)
                        pd_excel = pd.read_excel(f'{excel_path}/{e}')

                        lista_pvf = []
                        #Agrupar por puntos de venta           
                        for i, f in pd_excel.iterrows():
                            if i > 0:
                                encontrado = False
                                
                                #Obtenemos el ultimo por cada punto de venta (Usar clase PVF)
                                for pvf in lista_pvf:
                                    if pvf.punto_de_venta == f[2]:
                                        encontrado = True
                                        #actualizar
                                        pvf.cliente = c.id_cliente
                                        pvf.factura = f[1]
                                        pvf.fecha = f[0]
                                        pvf.n_factura = f[3]
                                        break
                                    
                                if not encontrado:
                                    #agregar
                                    new_pvf = PVF(c.id_cliente, f[1], f[0], f[2], f[3])
                                    lista_pvf.append(new_pvf)
                        
                                    
                        
                        if exists_excel:
                            lista_index = 0
                            for r in lista_pvf:
                                celda_pv = hoja_base.cell(row = c.linea, column = 54 + lista_index*3)
                                celda_fid = hoja_base.cell(row = c.linea, column = 55 + lista_index*3)
                                celda_f = hoja_base.cell(row = c.linea, column = 56 + lista_index*3)


                                #celda_pv = hoja_base[str(54 + lista_index*2) + str(c.linea)]
                                #celda_f = hoja_base[str(55 + lista_index*2) + str(c.linea)]

                                celda_pv.value = r.punto_de_venta
                                celda_fid.value = r.n_factura
                                celda_f.value = r.factura
                                
                                
                                lista_index = lista_index + 1
                    break
            else:
                print(f'No existe la carpeta del cliente {c.razon_social}')

            print(f'{c.razon_social}: ({index}/{len(lista_clientes)}) {int(index/len(lista_clientes)*100)}%')


        except FileNotFoundError:
            print(f'Error de {c.razon_social}: Archivo no encontrado')
                #celda_logs = hoja_logs[f'A{index}']
                #celda_logs.value = f'El cliente {c.cliente} no se pudo encontrar dentro de la carpeta WNS'
        except IndexError:
            print(f'Error de {c.razon_social}: Error en Archivo')
                #celda_logs = hoja_logs[f'A{index}']
                #celda_logs.value = f'El pdf del cliente {c.cliente} es incorrecto'
        except OSError:
            print(f'Error de {c.razon_social}: Ruta incorrecta')
                #celda_logs = hoja_logs[f'A{index}']
                #celda_logs.value = f'Error, la ruta es incorrecta para {c.cliente}'
    
        
    

    
    #Escribir en excel "Clientes a controlar" cada agrupación de PVF
else:
    print("La hoja no existe") 



openpy_doc.save(file_name)
openpy_doc.close()

    
#Fin